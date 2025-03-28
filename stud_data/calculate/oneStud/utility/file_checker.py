from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl

def formal(data):
	all_sub = set()
	
	for each in data:
		for item in data[each]:
			all_sub.add(item)
			
	all_sub = sorted(all_sub)
	
	new = {}
	for each in data:
		new[each] = {}
		for sub in all_sub:
			if sub in data[each]:
				new[each][sub] = data[each][sub]
				continue
			new[each][sub] = 0
	return new
			

def validate(file):
	
	keys = file.keys()
	
	for key in keys:
		needed = file[key]
		if type(needed) != dict:
			raise ValueError("each student mark should have subjects")
			
	for key_up in keys:
		needed = file[key_up]
		for key in needed.keys():
			try:
				float(needed[key])
			except ValueError:
				raise ValueError("Mark should be number only")
			except TypeError:
				raise ValueError("Mark should be number only")
		
	return formal(file)


def checkFile(input_data): 
	if type(input_data) == dict:
		return validate(input_data)
		
	file = Path(input_data)
	if not file.is_file():
		raise ValueError("the function receivs only dictionary or json file")
		
	suffix = file.suffix
		
	if suffix not in [".json",".xlsx"]:
		raise ValueError("the function receivs only dictionary or json file")
	
	if suffix == ".json":
		json = __import__("json")
		with open(input_data) as file:
			data = json.load(file)
	else:
		data = extractor(input_data)
		
	return validate(data)

	
def extractor(file):
	wb = load_workbook(file)
	
	ws = wb.active
	
	names = []
	row = 2
	ceil = f"A{row}"
	while ws[ceil].value != None:
		names.append(ws[ceil].value)
		row += 1
		ceil = f"A{row}"
		
		
	subs = []
	col = 2
	ceil = f"{gcl(col)}1"
	while ws[ceil].value != None:
		subs.append(ws[ceil].value)
		col += 1
		ceil = f"{gcl(col)}1"
		
	
	marks = []
	for each in range(2,len(names)+2):
		mark = []
		for sub in range(2,len(subs)+2):
			ceil = f"{gcl(sub)}{each}"
			value = ws[ceil].value
			mark.append(value)
			
		marks.append(mark)
		
	data = {}
	
	for i,name in enumerate(names):
		stud = {}
		for j,sub in enumerate(subs):
			stud[sub] = marks[i][j]
		data[name] = stud
		
	return data


def dict_xlsx(data,names,subject):	

	first = ["Name"] + subject
	
	all_stud = [first]
	
	one_sub = data[names[0]][subject[0]]
	if type(one_sub) == dict:
		keys = [""] + list(one_sub.keys()) * len(subject)
		all_stud.append(keys)
	for each in names:
		elem = [each]
		for item in data[each]:
			item = data[each][item]
			if type(item) == dict:
				
				for key in item:
					
					elem.append(item[key])
				continue
			elem.append(item)
		all_stud.append(elem)
	
	return all_stud
	
