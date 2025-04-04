"""This module has a class used to retrieve data from a given students data
has class:
        StudMark
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as gcl
from .one_stud.utility.file_checker import check_file, dict_xlsx,key_extractor,non_key_deleter
from .one_stud.grading import (
    grade_fun,
    gpa_fun,
    point_fun,
    stud_names,
    stud_marks,
    sub_list,
    max_min,
)


class StudMark:
    """This class has all operations that are related to students mark.
    has attributes:
            stud_data
            students
            subjects
            all_mark
            total
            min_mark
            max_mark

    has methods:
            sort
            grade
            point
            gpa
            average
            ranker
            id_giver
            one_sub
            one_stud
            json
            xlsx
    """

    def __init__(self, file):
        """initializes the following private attributes for the class:
        stud_data - the whole students data with dictionary
        students - list of students name
        subjects - list of subjects
        - all_mark - all marks of students in the form of 2D list
        - total - total number of students
        - max_sub - the subject that has maximum average mark from all students
        - min_sub - the subject that has minimum average mark from all students
        """
        self._stud_data = check_file(file)
        self._students = stud_names(self._stud_data)
        self._subjects = sub_list(self._stud_data)
        self._all_mark = stud_marks(self._stud_data)
        self._total = len(self._students)
        self._max_sub, self._min_sub = max_min(self._stud_data)

    @property
    def stud_data(self):
        """Used to access private _studData from outer scope"""
        return self._stud_data

    @property
    def students(self):
        """Used to access private _students from outer scope"""
        return self._students

    @property
    def subjects(self):
        """Used to access private _subjects from outer scope"""
        return self._subjects

    @property
    def all_mark(self):
        """Used to access private _all_mark from outer scope"""
        return self._all_mark

    @property
    def total(self):
        """Used to access private _total from outer scope"""
        return self._total

    @property
    def max_sub(self):
        """Used to access private _max_sub from outer scope"""
        return self._max_sub

    @property
    def min_sub(self):
        """Used to access private _min_sub from outer scope"""
        return self._min_sub

    def sort(self, gpa=False, update=False):
        """used to sort the students data, according to:
                - there gpa if gpa is in the dict and gpa parameter is True
                - else by there name
        attributes:
                - update - is bool to know if the user is need to update data or not.
                default is False
        """
        data = self._stud_data.copy()

        keys = list(data.keys())
        first = keys[0]

        if "gpa" in data[first] and gpa:
            data = {
                key: data[key]
                for key in sorted(data, key=lambda y: data[y]["gpa"], reverse=True)
            }

            return data

        data = {key: data[key] for key in sorted(data)}
        if update:
            self._stud_data = data.copy()
            return self._stud_data
        

        return data

    def grade(self, update=False):
        """Used to give grade for each student and for each subject
        - update - is bool to know if the user is need to update data or not.
                default is False
        """
        data = self._stud_data.copy()
        avl_key = key_extractor(data)

        subject = self._subjects.copy()

        for each in data:
            for sub in subject:
                if not isinstance(data[each][sub], dict):
                    mark = data[each][sub]
                    grade_sub = grade_fun(mark)

                    data[each][sub] = {"mark": mark, "grade": grade_sub}
                    continue
                mark = data[each][sub]["mark"]
                data[each][sub]["grade"] = grade_fun(mark)
        if update:
            self._stud_data = data.copy()
            return self._stud_data
        
        non_key_deleter(data,avl_key)

        return data

    def point(self, update=False):
        """Used to give point for each student and for each subject
        - update - is bool to know if the user is need to update data or not.
                default is False
        """
        data = self._stud_data.copy()
        avl_key = key_extractor(data)

        subject = self._subjects.copy()

        for each in data:
            for sub in subject:
                if not isinstance(data[each][sub], dict):
                    mark = data[each][sub]
                    point_sub = point_fun(mark)

                    data[each][sub] = {"mark": mark, "point": point_sub}
                    continue
                mark = data[each][sub]["mark"]
                data[each][sub]["point"] = point_fun(mark)
        if update:
            self._stud_data = data.copy()
            return self._stud_data
        
        non_key_deleter(data,avl_key)

        return data

    def gpa(self, cr_hours, update=False):
        """Used to calculate gpa by receiving credit hours
        for each subject and returns a dictionary with students data plus there gpa
        - update - is bool to know if the user is need to update data or not.
                default is False
        """
        data = self._stud_data.copy()
        avl_key = key_extractor(data)
        marks = self._all_mark

        gpa_list = [gpa_fun(each, cr_hours) for each in marks]

        for i, each in enumerate(self._students):
            data[each]["gpa"] = gpa_list[i]

        if update:
            self._stud_data = data.copy()
            return self._stud_data
        non_key_deleter(data,avl_key)

        return data

    def average(self, update=False):
        """Used to calculate average of students mark and
        returns students data dictionary plus there average
        - update - is bool to know if the user is need to update data or not.
                default is False
        """
        data = self._stud_data.copy()
        avl_key = key_extractor(data)

        num_sub = len(self._subjects)
        mark_list = self._all_mark

        aver = [round(sum(marks) / num_sub, 2) for marks in mark_list]
        for i, each in enumerate(self._students):
            data[each]["average"] = aver[i]
        if update:
            self._stud_data = data.copy()
            return self._stud_data
            
        non_key_deleter(data,avl_key)

        return data

    def ranker(self, gpa=False, update=False):
        """give rank for students by using:
                - gpa if gpa is in dict and gpa is True
                - else average
        - update - is bool to know if the user is need to update data or not.
                default is False
        """
        data = self._stud_data.copy()
        avl_key = key_extractor(data)
        
        students = self.students

        ave_key = "average" in data[students[0]]

        if gpa and "gpa" in data[students[0]]:
            sorted_data = self.sort(gpa)

            students_name = list(sorted_data.keys())

            rank = list(enumerate(students_name, 1))

        else:
            ave_data = self.average()

            rank = list(
                enumerate(
                    sorted(
                        ave_data, key=lambda y: ave_data[y]["average"], reverse=True
                    ),
                    1,
                )
            )

        for i, each in rank:
            data[each]["rank"] = i

        if update:
            self._stud_data = data.copy()
            return self._stud_data
            
        non_key_deleter(data,avl_key)

        return data

    def id_giver(self, pre="", suf="", sep="", length=4):
        """Used to give id for all students from starting 1.
        uses arguments:
                pre - to set prefix
                suf - to set suffix
                sep - to set separator between prefix and the id also between id and suffix
                length - length of id
        """
        data = self.sort()
        sep = sep if sep in ["-", "\\", ":", "/"] else ""

        def form(txt):
            return f"{pre}{sep}{txt:}{sep}{suf}"

        ids = [form(f"{i:0{length}d}") for i in range(1, len(data) + 1)]

        for i, each in enumerate(data):
            data[each]["id_number"] = ids[i]

        self._stud_data = data.copy()

    def one_sub(self, subject):
        """Used to retrieve information about single subject
        recives:
                subject - string of the subject
        returns :
                Subject class to retrieve data about the subject
        """
        data = self._stud_data.copy()
        marks = [data[name][subject] for name in data]

        class Subject:
            """Class used to retrieve different informations about the subject
            have the following methods:
                    max_mark
                    min_mark
                    average
                    full_mark
            """

            def max_mark(self, target=""):
                """this method used to get information about the largest
                mark in the list of all marks.
                recives:
                        target - string value must be 'sub' or 'stud' default is ''

                if target is 'sub':
                        returns maximum mark float type
                else:
                        if target is 'stud':
                                return list of students that has maximum mark
                        else:
                                return dict of students and mark with maximum mark

                """
                maximum = max(marks)
                result = {}
                for each in data:
                    value = data[each][subject]
                    if value == maximum:
                        result[each] = value

                if target.lower().strip() == "sub":
                    return maximum
                if target.lower().strip() == "stud":
                    studs = list(result.keys())
                    return studs

                return result

            def min_mark(self, target=""):
                """this method used to get information about the smalles
                 mark in the list of all marks
                recives:
                        target - string value must be 'sub' or 'stud' default is ''

                if target is 'sub':
                        returns minimum mark float type
                else:
                        if target is 'stud':
                                return list of students that has mimimum mark
                        else:
                                return dict of students and mark with minimum mark

                """
                minimum = min(marks)
                result = {}
                for each in data:
                    value = data[each][subject]
                    if value == minimum:
                        result[each] = value

                if target.lower().strip() == "Mark":
                    return minimum
                if target.lower().strip() == "stud":
                    studs = list(result.keys())
                    return studs

                return result

            def average(self):
                """used to calculate average mark of all students."""
                return round(sum(marks) / len(marks), 2)

            def full_mark(self):
                """used to extract dict
                with key of students and value of there mark associated with the given subject.
                """
                result = {name: data[name][subject] for name in data}
                return result

        return Subject()

    def one_stud(self, name):
        """Used to retrieve information about single student
        recives:
                name - string of the student's name
        returns :
                Student class to retrieve data about the student
        """
        data = self._stud_data.copy()
        stud_data = {name: data[name]}
        subjects = self._subjects.copy()

        class Student:
            """Class used to retrieve different informations about the student
            have the following attributes:
                    info
                    marks
            have the following methods:
                    max_mark
                    min_mark
                    average
                    grade
                    point
                    gpa
            """

            def __init__(self):
                """Used to initialize
                attributes for The class
                initializes:
                        info - dict with full information about the student
                        marks - list of marks the student have
                """
                self._info = stud_data.copy()
                self._marks = [stud_data[name][sub] for sub in stud_data[name]]
            
            @property
            def info(self):
            	return self._info
            
            @property
            def marks(self):
            	return self._marks

            def max_mark(self, target=""):
                """this method used to get information about the largest mark of the student.
                recives:
                        target - string value must be 'sub' or 'mark' default is ''

                if target is 'mark':
                        returns maximum mark float type
                else:
                        if target is 'sub':
                                return list of subjects that has maximum mark
                        else:
                                return dict of subjects and marks with maximum mark

                """
                maximum = max(self._marks)
                result = {}
                for each in stud_data[name]:
                    value = data[name][each]
                    if value == maximum:
                        result[each] = value
                target = target.lower().strip()
                if target == "mark":
                    return maximum
                if target == "sub":
                    return list(result.keys())
                return result

            def min_mark(self, target=""):
                """this method used to get information about the smallest mark of the student.
                recives:
                        target - string value must be 'sub' or 'mark' default is ''

                if target is 'mark':
                        returns minimum mark float type
                else:
                        if target is 'sub':
                                return list of subjects that has minimum mark
                        else:
                                return dict of subjects and marks with minimum mark

                """
                minimum = min(self._marks)
                result = {}
                for each in stud_data[name]:
                    value = data[name][each]
                    if value == minimum:
                        result[each] = value
                target = target.lower().strip()
                if target == "mark":
                    return minimum
                if target == "sub":
                    return list(result.keys())
                return result

            def average(self):
                """used to get average mark of the student"""
                return sum(self._marks) / len(self._marks)

            def grade(self, sub="all"):
                """used to get grades of the student:
                recives:
                        sub - string with possible subjects default is 'all'
                if subject is given:
                        return grade of given subject
                else:
                        return all grade
                """
                if sub in subjects:
                    mark = stud_data[name][sub]
                    return grade_fun(mark)
                grades = [grade_fun(each) for each in self._marks]

                return grades

            def point(self, sub="all"):
                """used to get point of the student:
                recives:
                        sub - string with possible subjects default is 'all'
                if subject is given:
                        return point of given subject
                else:
                        return all point
                """
                if sub in subjects:
                    mark = stud_data[name][sub]
                    return point_fun(mark)

                point = [point_fun(each) for each in self._marks]

                return point

            def gpa(self, cr_hours=None):
                """used to calculate gpa of the student
                recive:
                        cr_hours - list or tuple of credit hours

                if cr_hours doesn't given:
                        check if a student has gpa in the dict
                        if it has:
                                return gpa
                        else:
                                raise TypeError
                else:
                        call gpaFun for calculating gpa and return it
                """
                try:
                    return gpa_fun(self._marks, cr_hours)
                except (TypeError, ValueError):
                    if "gpa" in stud_data[name]:
                        return stud_data[name]["gpa"]
                    raise TypeError(
                        "gpa() missing 1 required positional argument: 'cr_hours'"
                    )

        return Student()

    def json(self):
        """used to convert dictionary data to json file."""
        json = __import__("json")
        with open("student_data.json", "w") as file:
            json.dump(self._stud_data, file, indent=4)

        print("student data has been saved")

    def xlsx(self):
        """Used to change dictionary data to xlsx format."""
        data = self._stud_data
        rowork_sheet = dict_xlsx(self._stud_data, self._students, self._subjects)
        work_book = Workbook()

        work_sheet = work_book.active
        work_sheet.title = "Student_data"

        for each in rowork_sheet:

            work_sheet.append(each)
        in_1 = self._students[0]

        mg_strt = 2
        for each in data[in_1]:
            each = data[in_1][each]
            if isinstance(each, dict):
                mg_end = mg_strt + len(each) - 1
                mv_strt = mg_strt + 1
                mv_end = mg_strt + len(self._subjects) - 1
                ch1 = f"{gcl(mg_strt)}1"
                ch2 = f"{gcl(mg_end)}1"
                ch3 = f"{gcl(mv_strt)}1"
                ch4 = f"{gcl(mv_end)}1"

                work_sheet.move_range(f"{ch3}:{ch4}", cols=len(each) - 1)
                work_sheet.merge_cells(f"{ch1}:{ch2}")

                mg_strt += len(each)

        work_book.save("student_info.xlsx")
