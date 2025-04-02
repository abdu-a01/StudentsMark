"""This module provides functions for validating input types and
extracting data from EXEL and json files.
The module have the following functions:
        -formal
        -validate
        -checkFile
        -extractor
        -dict_xlsx
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl


def formal(data):
    """This function recives student data with dictionary form and
    then returns dictionary which has students with the same subjects
    if a student doesn't has a subject:
            give the subject as key with value 0
    else:
            it keeps the previous state
    """
    all_sub = set()

    for each in data:
        for item in data[each]:
            all_sub.add(item)

    all_sub = sorted(all_sub)

    new = {}
    for each in data:
        new[each] = {}
        for sub in all_sub:
            if sub in data[each]:
                new[each][sub] = data[each][sub]
                continue
            new[each][sub] = 0
    return new


def validate(file):
    """this function recives dictionary and check if it has proper marks and subjects or not
    if it has:
            call formal function
    ekse:
            raise ValueError
    """

    keys = file.keys()

    for key in keys:
        needed = file[key]
        if not isinstance(needed, dict):
            raise ValueError("each student mark should have subjects and marks")

    for key_up in keys:
        needed = file[key_up]
        for key in needed.keys():
            try:
                float(needed[key])
            except (ValueError, TypeError):
                raise ValueError("Mark should be number only")

    return formal(file)


def check_file(input_data):
    """This function recives data and check the type of the data.
    if it is dictionary:
            call validate function
    if it is json file :
            dump it to dict then call validate function
    if it is xlsx file:
            call extractor function
            then call validate function
    else:
            raise ValueError
    """
    if isinstance(input_data, dict):
        return validate(input_data)

    file = Path(input_data)
    if not file.is_file():
        raise ValueError("the function receivs only dictionary or json file")

    suffix = file.suffix

    if suffix not in [".json", ".xlsx"]:
        raise ValueError("the function receivs only dictionary or json file")

    if suffix == ".json":
        json = __import__("json")
        with open(input_data) as file:
            data = json.load(file)
    else:
        data = extractor(input_data)

    return validate(data)


def extractor(file):
    """This function recives xlsx file and extact the data in to dictionary:
    return dict
    """
    work_book = load_workbook(file)

    work_sheet = work_book.active
    work_sheet.title = "Students"

    names = []
    row = 2

    while work_sheet[f"A{row}"].value is not None:
        names.append(work_sheet[f"A{row}"].value)
        row += 1

    subs = []
    col = 2
    while work_sheet[f"{gcl(col)}1"].value is not None:
        subs.append(work_sheet[f"{gcl(col)}1"].value)
        col += 1

    marks = []
    for _ in range(2, len(names) + 2):
        mark = [float(work_sheet[f"{gcl(col)}1"].value) for _ in range(2, len(subs) + 2)]

        marks.append(mark)

    data = {}

    for i, name in enumerate(names):
        stud = {sub: marks[i][j] for j, sub in enumerate(subs)}

        data[name] = stud

    return data


def dict_xlsx(data, names, subject):
    """This function recives:
            - data - dictionary with students data
            - names - list of students name
            - subject - list of subjects
    return list of data for each xlsx rows
    """

    first = ["Name"] + subject

    all_stud = [first]

    one_sub = data[names[0]][subject[0]]
    if isinstance(one_sub, dict):
        keys = [""] + list(one_sub.keys()) * len(subject)
        all_stud.append(keys)
    for each in names:
        elem = [each]
        for item in data[each]:
            item = data[each][item]
            if isinstance(item, dict):

                for key in item:

                    elem.append(item[key])
                continue
            elem.append(item)
        all_stud.append(elem)

    return all_stud
